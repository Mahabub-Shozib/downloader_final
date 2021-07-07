import openpyxl
import youtube_dl

excel = openpyxl.load_workbook("2.xlsx")
sheets =excel.sheetnames
active = excel.active.title
print(active)
# sheetName = str(input("Please enter the sheet name: "))
sh1 = excel['download']
max_row = sh1.max_row
max_col = sh1.max_column

#TODO for extract all link from a playlist
def extract_link(playlistUrl):
    youtube_download_option = {}
    ydl = youtube_dl.YoutubeDL(youtube_download_option)
    video = ""
    with ydl:
        result = ydl.extract_info(playlistUrl, download=False)

        if 'entries' in result:
            # Can be a playlist or a list of videos
            video = result['entries']

            # loops entries to grab each video_url
            for i, item in enumerate(video):
                video = result['entries'][i]['webpage_url']
                print(video)
                ydl.download(video)


def total_videos_of_single_playlist():


for row in range(1,max_row+1):
    for col in range(2,max_col+1):
        playlist_link = sh1.cell(row,col).value
        print(playlist_link)
        extract_link(playlist_link)


def write_playlist_name():


